from selenium import webdriver
import openpyxl
import random
import time

characters = ['a', 'A', 'b', 'B', 'c', 'C', 'd', 'D', 'e', 'E', 'f', 'F', 'g', 'G', 'h', 'H', 'i', 'I',
              'j', 'J', 'k', 'K', 'l', 'L', 'm', 'M', 'n', 'N', 'o', 'O', 'p', 'P', 'q', 'Q', 'r', 'R',
              's', 'S', 't', 'T', 'u', 'U', 'v', 'V', 'w', 'W', 'x', 'X', 'y', 'Y', 'z', 'Z', '0', '1',
              '2', '3', '4', '5', '6', '7', '9', '!', '@', "#", "$", "%", "^", "&", '*', "(", ")"]
domains = ['@gmail.com', '@yahoo.com', '@hotmail.com', '@outlook.com', '@inbox.com', '@mail.com', '@aol.com',
           '@icloud.com', '@zoho.com', '@yandex.com', '@protonmail.com']


def generate_random_name():
    wb = openpyxl.load_workbook('names.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    rand_int_first = random.randint(1, sheet.max_row)
    rand_int_middle = random.randint(1, sheet.max_row)
    rand_int_last = random.randint(1, sheet.max_row)
    first_name = sheet.cell(row=rand_int_first, column=1).value
    middle_name = sheet.cell(row=rand_int_middle, column=1).value
    last_name = sheet.cell(row=rand_int_last, column=1).value
    return [first_name, middle_name, last_name]


def generate_random_email(domains, first, middle, last):
    rand_int = random.randint(0, len(domains) - 1)
    email = last.lower() + '_' + middle.lower() + '_' + first.lower() + str(random.randint(100, 1000)) + domains[rand_int]
    return email


def generate_random_password(characters):
    length = random.randint(6, 11)
    password = ''
    for i in range(length):
        rand_int = random.randint(0, len(characters) - 1)
        password += characters[rand_int]
    return password


def name_list_to_string(list):
    string = ''
    for i in range(len(list)):
        string += list[i] + ' '
    return string


if __name__ == '__main__':
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--incognito")
    driver = webdriver.Chrome(chrome_options=chrome_options)
    driver.implicitly_wait(10)
    driver.get('https://www.instagram.com/')
    seen = False

    while True:
        random_person = generate_random_name()
        random_person_name = name_list_to_string(random_person)
        print(random_person_name)
        random_email = generate_random_email(domains, random_person[0], random_person[1], random_person[2])
        print(random_email)
        random_password = generate_random_password(characters)
        print(random_password)

        driver.find_element_by_name('emailOrPhone').send_keys(random_email)
        driver.find_element_by_name('fullName').send_keys(random_person_name)
        driver.find_element_by_xpath('//*[@id="react-root"]/section/main/article/div[2]/div[1]/div/form/div[5]/div/div[2]/div/button/span').click()
        random_username = driver.find_element_by_name('username').get_attribute('value')
        print(random_username)
        driver.find_element_by_name('password').send_keys(random_password)
        driver.find_element_by_xpath('//*[@id="react-root"]/section/main/article/div[2]/div[1]/div/form/div[7]/div/button').click()

        if not seen:
            driver.find_element_by_xpath('/html/body/div[3]/div/div/div[3]/button[2]').click()
            seen = True

        driver.find_element_by_xpath('//*[@id="react-root"]/section/nav/div[2]/div/div/div[2]/input').send_keys('anklebreakervo')
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="react-root"]/section/nav/div[2]/div/div/div[2]/div[2]/div[2]/div/a[1]/div').click()
        driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/header/section/div[1]/button').click()
        driver.find_element_by_xpath('//*[@id="react-root"]/section/nav/div[2]/div/div/div[3]/div/div[3]/a/span').click()
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/header/section/div[1]/div/button/span').click()
        time.sleep(1)
        driver.find_element_by_xpath('/html/body/div[3]/div/div/div/button[6]').click()

        wb = openpyxl.load_workbook('accounts.xlsx')
        sheet = wb.active
        current_row = sheet.max_row + 1
        sheet.cell(row=current_row, column=1).value = random_email
        sheet.cell(row=current_row, column=2).value = random_person_name
        sheet.cell(row=current_row, column=3).value = random_username
        sheet.cell(row=current_row, column=4).value = random_password
        wb.save('accounts.xlsx')
